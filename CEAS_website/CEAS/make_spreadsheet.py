from CEAS.models import *
from openpyxl import Workbook
from openpyxl.styles import Color, Fill, PatternFill

def make_spreadsheet():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Matrix" # title of the tab in Excel

    # Lists of headers used
    student_headers = ["Code:\n1-recommend\n2-possibility\n3-don't recommend-GPA\n4-Dec grad\n5-Disqualified\n6-All choices not open\n7-withdrew", "Name", "Primary Major", "Secondary Major", "Major on system if different", "Level in school", "Grad Date", "Self-Reported GPA", "GPA after spring semester", "Gender", "Ethnicity", "Previous Research Experience", "Other employment plans", "Applied before?"]
    project_headers = ["Department", "Faculty", "AltFaculty", "EDC Focus", "Semester", "Special Requirements", "StudName", "Speedtype", "Donor?", "# Applicants", "# Selected", "Name / Choice Given",]
    end_student_headers = ["Comment", "Skill1", "Skill2", "Skill3", "Project1", "Project2", "Project3", "Project4", "Project5", "Student ID", "Boulder Email", "Summer Email", "Other Employment"]

    # Import all objects
    project_list = Project.objects.all()
    student_list = Student.objects.all()

    # set up headers for spreadsheet
    for row in range(1, len(project_headers) + 1):
        _ = ws1.cell(column=len(student_headers) + 1, row=row, value=project_headers[row - 1])
    for col in range(1, len(student_headers) + 1):
        _ = ws1.cell(column=col, row=len(project_headers), value=student_headers[col - 1])
    for col in range(1, len(end_student_headers) + 1):
        _ = ws1.cell(column=col + len(student_headers) + len(project_list) + 1, row=len(project_headers), value=end_student_headers[col - 1])

    i = 0
    # Fill out student info
    i = 0
    for s in student_list:
        i += 1
        _ = ws1.cell(column=2, row=len(project_headers)+i, value=s.full_Name)
        _ = ws1.cell(column=3, row=len(project_headers)+i, value=s.primary_Major)
        _ = ws1.cell(column=4, row=len(project_headers)+i, value=s.secondary_Major)
        _ = ws1.cell(column=6, row=len(project_headers)+i, value=s.grade_level_as_of_next_fall)
        _ = ws1.cell(column=7, row=len(project_headers)+i, value=s.anticipated_Graduation_Date)
        _ = ws1.cell(column=8, row=len(project_headers)+i, value=float(s.GPA))
        _ = ws1.cell(column=10, row=len(project_headers)+i, value=s.gender)
        _ = ws1.cell(column=11, row=len(project_headers)+i, value=s.race)
        _ = ws1.cell(column=12, row=len(project_headers)+i, value=s.prevres)
        _ = ws1.cell(column=14, row=len(project_headers)+i, value=s.previously_Applied)
        # this is where the project matches are added to the matrix
        _ = ws1.cell(column=15 + int(s.fifthChoice), row=len(project_headers)+i, value=5)
        _ = ws1.cell(column=15 + int(s.fourthChoice), row=len(project_headers)+i, value=4)
        _ = ws1.cell(column=15 + int(s.thirdChoice), row=len(project_headers)+i, value=3)
        _ = ws1.cell(column=15 + int(s.secondChoice), row=len(project_headers)+i, value=2)
        _ = ws1.cell(column=15 + int(s.firstChoice), row=len(project_headers)+i, value=1)
        # these are the columns on the far right
        _ = ws1.cell(column=len(student_headers)+len(project_list) + 3, row=len(project_headers)+i, value=s.skill1)
        _ = ws1.cell(column=len(student_headers)+len(project_list) + 4, row=len(project_headers)+i, value=s.skill2)
        _ = ws1.cell(column=len(student_headers)+len(project_list) + 5, row=len(project_headers)+i, value=s.skill3)
        _ = ws1.cell(column=len(student_headers)+len(project_list) + 11, row=len(project_headers)+i, value=int(s.student_ID))
        _ = ws1.cell(column=len(student_headers)+len(project_list) + 12, row=len(project_headers)+i, value=s.email)
        _ = ws1.cell(column=len(student_headers)+len(project_list) + 14, row=len(project_headers)+i, value=s.feplans)

    # Fill out project info
    j = 0
    for p in project_list:
        j += 1
        _ = ws1.cell(column=len(student_headers)+1+j, row=1, value=p.primary_faculty_department)
        _ = ws1.cell(column=len(student_headers)+1+j, row=2, value=p.primary_faculty_member)
        _ = ws1.cell(column=len(student_headers)+1+j, row=3, value=p.secondary_faculty_member)
        _ = ws1.cell(column=len(student_headers)+1+j, row=5, value="BOTH")
        _ = ws1.cell(column=len(student_headers)+1+j, row=6, value=p.special_requirements)
        _ = ws1.cell(column=len(student_headers)+1+j, row=7, value=p.grad_post_doc_info)
        _ = ws1.cell(column=len(student_headers)+1+j, row=8, value=p.speed_type)
        _ = ws1.cell(column=len(student_headers)+1+j, row=10, value=p.number_of_applicants)
        _ = ws1.cell(column=len(student_headers)+1+j, row=12, value=p.title)

    # save file
    wb.save("static/Matrix.xlsx")

    return
